#!/usr/bin/env python3
"""
RevRating - InfoReady Review Processor
=======================================
Cross-platform replacement for the RevRatingUMD VBA macro.
Reads any InfoReady Application Reviews export and produces
a formatted MacroResult .xlsx workbook.

Works with any competition type (limited submissions, internal
funding, fellowships, etc.) as long as the export contains
the standard InfoReady Reviewer Details columns.

Usage:
    python3 rev_rating.py "Application Reviews Export.xlsx"
    python3 rev_rating.py "my_export.xlsb"
    python3 rev_rating.py "my_export.csv"

Output:
    MacroResult_YYYY-MM-DD_HH-MM.xlsx  (saved in same folder as input)

Requires:
    pip3 install openpyxl pyxlsb
"""

import sys
import os
import csv
from datetime import datetime
from collections import OrderedDict, defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


# ============================================================
# COLUMN NAME ALIASES
# ============================================================
# InfoReady exports may vary slightly in column naming across
# different competition types. These aliases handle known
# variations so the script works without manual column renaming.

COLUMN_ALIASES = {
    "application": ["application", "application title", "app title", "proposal", "proposal title"],
    "applicant": ["applicant", "applicant name", "pi", "pi name", "principal investigator"],
    "reviewer": ["reviewer", "reviewer name", "panelist", "evaluator"],
    "label": ["comment label", "label", "rubric category", "criterion", "criteria", "category"],
    "rating": ["rating", "score", "numeric rating", "numeric score"],
    "comment": ["comments to applicant", "comment to applicant", "applicant comments",
                 "reviewer comments", "comments", "feedback"],
    "admin": ["administrator comments", "admin comments", "administrator comment",
              "admin comment", "internal comments", "internal notes"],
}


# ============================================================
# DATA READING
# ============================================================

def read_reviewer_details(filepath):
    """Read Reviewer Details from .xlsx, .xlsb, or .csv and return (headers, rows)."""
    ext = Path(filepath).suffix.lower()

    if ext == ".xlsb":
        return _read_xlsb(filepath)
    elif ext == ".csv":
        return _read_csv(filepath)
    elif ext in (".xlsx", ".xls"):
        return _read_xlsx(filepath)
    else:
        try:
            return _read_xlsx(filepath)
        except Exception:
            return _read_csv(filepath)


def _read_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = _find_reviewer_sheet_xlsx(wb)

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) == 0:
        return [], []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    data = rows[1:]
    return headers, data


def _find_reviewer_sheet_xlsx(wb):
    """Find the best sheet to use. Priority: Reviewer Details > any sheet with 'review' > active."""
    # Exact match first
    for name in wb.sheetnames:
        if "reviewer" in name.lower() and "detail" in name.lower():
            return wb[name]
    # Partial match
    for name in wb.sheetnames:
        if "review" in name.lower():
            return wb[name]
    return wb.active


def _read_xlsb(filepath):
    try:
        from pyxlsb import open_workbook
    except ImportError:
        print("ERROR: pyxlsb not installed. Run: pip3 install pyxlsb")
        sys.exit(1)

    wb = open_workbook(filepath)
    sheet_name = _find_reviewer_sheet_xlsb(wb)

    if sheet_name is None:
        return [], []

    rows = []
    with wb.get_sheet(sheet_name) as sheet:
        for row in sheet.rows():
            rows.append([cell.v for cell in row])

    if len(rows) == 0:
        return [], []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    data = rows[1:]
    return headers, data


def _find_reviewer_sheet_xlsb(wb):
    """Find best sheet in xlsb workbook."""
    for name in wb.sheets:
        if "reviewer" in name.lower() and "detail" in name.lower():
            return name
    for name in wb.sheets:
        if "review" in name.lower():
            return name
    return wb.sheets[-1] if wb.sheets else None


def _read_csv(filepath):
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if len(rows) == 0:
        return [], []
    headers = [h.strip() for h in rows[0]]
    data = rows[1:]
    return headers, data


def col_index(headers, field_name):
    """Find column index using aliases. Returns -1 if not found."""
    aliases = COLUMN_ALIASES.get(field_name, [field_name])
    headers_lower = [h.lower().strip() for h in headers]
    for alias in aliases:
        for i, h in enumerate(headers_lower):
            if h == alias:
                return i
    # Fallback: partial match (header contains alias or alias contains header)
    for alias in aliases:
        for i, h in enumerate(headers_lower):
            if alias in h or h in alias:
                if len(h) > 3:  # avoid matching tiny strings
                    return i
    return -1


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def safe_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ============================================================
# DATA PROCESSING
# ============================================================

def process_data(headers, data):
    """Process raw data into structured dictionaries for all sheets."""
    ci = {
        "app": col_index(headers, "application"),
        "applicant": col_index(headers, "applicant"),
        "reviewer": col_index(headers, "reviewer"),
        "label": col_index(headers, "label"),
        "rating": col_index(headers, "rating"),
        "comment": col_index(headers, "comment"),
        "admin": col_index(headers, "admin"),
    }

    # Report what was found
    print("  Column mapping:")
    friendly = {
        "app": "Application", "applicant": "Applicant", "reviewer": "Reviewer",
        "label": "Comment Label", "rating": "Rating",
        "comment": "Comments to Applicant", "admin": "Administrator Comments"
    }
    for key, idx in ci.items():
        status = f"column {idx} ({headers[idx]})" if idx >= 0 else "NOT FOUND"
        req = " [REQUIRED]" if key in ("app", "reviewer", "rating") else ""
        print(f"    {friendly[key]}: {status}{req}")

    if ci["app"] < 0 or ci["reviewer"] < 0 or ci["rating"] < 0:
        print("\nERROR: Required columns not found. Need: Application, Reviewer, Rating")
        print("  Tip: Check that the input file contains the Reviewer Details sheet")
        print(f"  Found headers: {headers}")
        sys.exit(1)

    # Ordered tracking
    apps = OrderedDict()
    app_applicant = {}
    reviewers = OrderedDict()
    scores = defaultdict(lambda: defaultdict(float))
    cat_ratings = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    labels = OrderedDict()
    comments = defaultdict(lambda: defaultdict(list))
    admin_comments = defaultdict(dict)
    reviewers_by_app = defaultdict(OrderedDict)

    has_labels = ci["label"] >= 0
    has_comments = ci["comment"] >= 0
    has_admin = ci["admin"] >= 0

    for row in data:
        def g(key):
            idx = ci[key]
            if idx < 0 or idx >= len(row):
                return ""
            return safe_str(row[idx])

        app = g("app")
        applicant = g("applicant")
        reviewer = g("reviewer")
        label = g("label") if has_labels else ""
        rating = safe_float(row[ci["rating"]] if ci["rating"] < len(row) else None)
        comment = g("comment") if has_comments else ""
        admin = g("admin") if has_admin else ""

        if not app or not reviewer:
            continue

        # Track apps and applicants
        if app not in apps:
            apps[app] = len(apps) + 1
        if applicant and app not in app_applicant:
            app_applicant[app] = applicant

        # Track reviewers
        if reviewer not in reviewers:
            reviewers[reviewer] = len(reviewers) + 1

        # Track reviewers per app
        if reviewer not in reviewers_by_app[app]:
            reviewers_by_app[app][reviewer] = True

        # Admin comments (first one per reviewer+app)
        if has_admin and admin and app not in admin_comments[reviewer]:
            admin_comments[reviewer][app] = admin

        # Skip non-rated rows for score calculations
        if has_labels and not label:
            continue
        if rating is None:
            continue

        # Track labels
        if label and label not in labels:
            labels[label] = len(labels) + 1

        # Sum scores
        scores[app][reviewer] += rating

        # Category ratings
        if label:
            cat_ratings[app][reviewer][label] = rating

        # Comments
        if comment and comment.strip():
            comments[app][reviewer].append((label if label else "General", comment))

    return {
        "apps": apps,
        "app_applicant": app_applicant,
        "reviewers": reviewers,
        "scores": scores,
        "cat_ratings": cat_ratings,
        "labels": labels,
        "comments": comments,
        "admin_comments": admin_comments,
        "reviewers_by_app": reviewers_by_app,
        "has_labels": has_labels,
        "has_comments": has_comments,
        "has_admin": has_admin,
    }


def make_app_key(idx, app_title, applicant):
    """Create short app key like '01. IRG: Advancing Ceram_Ren'"""
    short = app_title[:20].strip()
    if applicant:
        last_name = applicant.split(",")[0].strip() if "," in applicant else applicant.strip()
    else:
        last_name = ""
    return f"{idx:02d}. {short}_{last_name}" if last_name else f"{idx:02d}. {short}"


# ============================================================
# STYLES
# ============================================================

BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
WHITE_FONT = Font(name="Calibri", size=11, color="FFFFFF")
WHITE_FONT_BOLD = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
BLACK_FONT = Font(name="Calibri", size=11, color="000000")
BLACK_FONT_BOLD = Font(name="Calibri", size=11, color="000000", bold=True)
RED_FONT_BOLD = Font(name="Calibri", size=11, color="FF0000", bold=True)
GRAY_FILL = PatternFill(start_color="E1E1E1", end_color="E1E1E1", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="B2B2B2", end_color="B2B2B2", fill_type="solid")
LEGEND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def apply_header_row(ws, row_num, max_col, bold=False):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = BLACK_FILL
        cell.font = WHITE_FONT_BOLD if bold else WHITE_FONT
        cell.alignment = CENTER_WRAP
        cell.border = THIN_BORDER


def apply_borders(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def apply_center_all(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = CENTER_WRAP
            cell.font = Font(name="Calibri", size=11)


# ============================================================
# SHEET BUILDERS
# ============================================================

def build_matrix_final(wb, d):
    """Build Matrix Final sheet with scores, stats, and normalized scores."""
    ws = wb["Matrix Final"]

    reviewer_list = sorted(d["reviewers"].keys(), key=str.lower)
    app_list = list(d["apps"].keys())

    if not app_list or not reviewer_list:
        ws.cell(row=1, column=1, value="No data to display")
        return None

    # Header rows
    ws.cell(row=1, column=1, value="Sum of Rating")
    ws.cell(row=2, column=1, value="Applicant")

    first_rc = 2
    for i, rev in enumerate(reviewer_list):
        ws.cell(row=2, column=first_rc + i, value=rev)
    last_rc = first_rc + len(reviewer_list) - 1

    col_gt = last_rc + 1
    col_av = col_gt + 1
    col_sd = col_av + 1
    col_n = col_sd + 1
    col_nm = col_n + 1

    ws.cell(row=2, column=col_gt, value="Grand Total")
    ws.cell(row=2, column=col_av, value="Average")
    ws.cell(row=2, column=col_sd, value="Standard Deviation")
    ws.cell(row=2, column=col_n, value="Number of Reviews")
    ws.cell(row=2, column=col_nm, value="Normalized Score")

    first_data_row = 3
    out_row = first_data_row

    for idx, app in enumerate(app_list, 1):
        applicant = d["app_applicant"].get(app, "")
        ws.cell(row=out_row, column=1, value=make_app_key(idx, app, applicant))

        row_scores = []
        for i, rev in enumerate(reviewer_list):
            val = d["scores"].get(app, {}).get(rev, None)
            col = first_rc + i
            if val is not None and val > 0:
                ws.cell(row=out_row, column=col, value=val)
                row_scores.append(val)
            else:
                ws.cell(row=out_row, column=col, value="")

        if row_scores:
            ws.cell(row=out_row, column=col_gt, value=sum(row_scores))
            ws.cell(row=out_row, column=col_av, value=sum(row_scores) / len(row_scores))
            mean = sum(row_scores) / len(row_scores)
            variance = sum((x - mean) ** 2 for x in row_scores) / len(row_scores)
            ws.cell(row=out_row, column=col_sd, value=variance ** 0.5)
            ws.cell(row=out_row, column=col_n, value=len(row_scores))

        out_row += 1

    last_data_row = out_row - 1

    row_gt = out_row
    row_av = row_gt + 1
    row_sd = row_gt + 2
    row_n = row_gt + 3
    row_bias = row_n + 1

    ws.cell(row=row_gt, column=1, value="Grand Total")
    ws.cell(row=row_av, column=1, value="Average")
    ws.cell(row=row_sd, column=1, value="Standard Deviation")
    ws.cell(row=row_n, column=1, value="Number of Reviews")
    ws.cell(row=row_bias, column=1, value="Deviation From Mean")

    for i, rev in enumerate(reviewer_list):
        col = first_rc + i
        col_vals = []
        for r in range(first_data_row, last_data_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None and v != "":
                col_vals.append(float(v))

        if col_vals:
            ws.cell(row=row_gt, column=col, value=sum(col_vals))
            ws.cell(row=row_av, column=col, value=sum(col_vals) / len(col_vals))
            mean_c = sum(col_vals) / len(col_vals)
            var_c = sum((x - mean_c) ** 2 for x in col_vals) / len(col_vals)
            ws.cell(row=row_sd, column=col, value=var_c ** 0.5)
            ws.cell(row=row_n, column=col, value=len(col_vals))

    # Bias
    for i, rev in enumerate(reviewer_list):
        col = first_rc + i
        deviations = []
        for r in range(first_data_row, last_data_row + 1):
            rev_score = ws.cell(row=r, column=col).value
            avg_score = ws.cell(row=r, column=col_av).value
            if rev_score is not None and rev_score != "" and avg_score is not None:
                deviations.append(float(rev_score) - float(avg_score))
        if deviations:
            ws.cell(row=row_bias, column=col, value=sum(deviations) / len(deviations))

    # Normalized Score
    for r in range(first_data_row, last_data_row + 1):
        adjusted = []
        for i, rev in enumerate(reviewer_list):
            col = first_rc + i
            rev_score = ws.cell(row=r, column=col).value
            bias = ws.cell(row=row_bias, column=col).value
            if rev_score is not None and rev_score != "" and bias is not None:
                adjusted.append(float(rev_score) - float(bias))
        if adjusted:
            ws.cell(row=r, column=col_nm, value=sum(adjusted) / len(adjusted))

    # Sort by Normalized Score descending
    data_rows = []
    for r in range(first_data_row, last_data_row + 1):
        row_data = []
        for c in range(1, col_nm + 1):
            row_data.append(ws.cell(row=r, column=c).value)
        norm = ws.cell(row=r, column=col_nm).value
        data_rows.append((norm if norm is not None else -9999, row_data))

    data_rows.sort(key=lambda x: x[0], reverse=True)

    for i, (_, row_data) in enumerate(data_rows):
        r = first_data_row + i
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # ---- FORMATTING ----
    max_col = col_nm
    ws.row_dimensions[1].hidden = True
    ws.row_dimensions[2].height = 45
    for r in range(first_data_row, row_bias + 1):
        ws.row_dimensions[r].height = 45

    ws.column_dimensions["A"].width = 32
    for c in range(first_rc, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    apply_header_row(ws, 2, max_col)
    apply_center_all(ws, 1, row_bias, 1, max_col)
    apply_borders(ws, 2, row_bias, 1, max_col)

    for r in range(first_data_row, row_bias + 1):
        for c in range(first_rc, col_gt + 1):
            ws.cell(row=r, column=c).number_format = "0"
        ws.cell(row=r, column=col_av).number_format = "0.00"
        ws.cell(row=r, column=col_sd).number_format = "0.00"
        ws.cell(row=r, column=col_n).number_format = "0"
        ws.cell(row=r, column=col_nm).number_format = "0.00"

    for c in range(first_rc, max_col + 1):
        ws.cell(row=row_bias, column=c).number_format = "0.00"

    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_gt, column=c)
        cell.fill = BLACK_FILL
        cell.font = WHITE_FONT

    sd_range = f"{get_column_letter(col_sd)}{first_data_row}:{get_column_letter(col_sd)}{last_data_row}"
    ws.conditional_formatting.add(sd_range,
        CellIsRule(operator="greaterThan", formula=["4"],
                   fill=LIGHT_RED_FILL, font=BLACK_FONT))

    return {
        "reviewer_list": reviewer_list,
        "first_rc": first_rc, "last_rc": last_rc,
        "col_gt": col_gt, "col_av": col_av, "col_sd": col_sd,
        "col_n": col_n, "col_nm": col_nm,
        "first_data_row": first_data_row, "last_data_row": last_data_row,
        "row_gt": row_gt, "row_av": row_av, "row_sd": row_sd,
        "row_n": row_n, "row_bias": row_bias,
    }


def build_reviewer_summary(wb, mf_info):
    """Build Reviewer Summary from Matrix Final data."""
    ws_mf = wb["Matrix Final"]
    ws = wb["Reviewer Summary"]

    if mf_info is None:
        ws.cell(row=1, column=1, value="No data to display")
        return

    headers = ["Reviewer", "Average", "Standard Deviation", "Number of Reviews", "Estimated Bias"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    rows = []
    for i, rev in enumerate(mf_info["reviewer_list"]):
        col = mf_info["first_rc"] + i
        avg = ws_mf.cell(row=mf_info["row_av"], column=col).value
        sd = ws_mf.cell(row=mf_info["row_sd"], column=col).value
        n = ws_mf.cell(row=mf_info["row_n"], column=col).value
        bias = ws_mf.cell(row=mf_info["row_bias"], column=col).value
        rows.append((rev, avg, sd, n, bias, bias if bias is not None else -9999))

    rows.sort(key=lambda x: x[5], reverse=True)

    for i, (rev, avg, sd, n, bias, _) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=rev)
        ws.cell(row=i, column=2, value=avg)
        ws.cell(row=i, column=3, value=sd)
        ws.cell(row=i, column=4, value=n)
        ws.cell(row=i, column=5, value=bias)

    last_r = 1 + len(rows)

    # Formatting
    apply_header_row(ws, 1, 5, bold=True)
    for r in range(1, last_r + 1):
        ws.row_dimensions[r].height = 40

    ws.column_dimensions["A"].width = 23
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 23

    apply_center_all(ws, 1, last_r, 1, 5)
    apply_borders(ws, 2, last_r, 1, 5)

    for r in range(2, last_r + 1):
        ws.cell(row=r, column=2).number_format = "0.00"
        ws.cell(row=r, column=3).number_format = "0.00"
        ws.cell(row=r, column=4).number_format = "0"
        ws.cell(row=r, column=5).number_format = "0.00"

    # Legend
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 48
    for r in range(1, 7):
        for c in range(7, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = LEGEND_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Calibri", size=11)

    ws.cell(row=1, column=7, value="Bias Legend").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=2, column=7, value="Estimated Bias < 0:")
    ws.cell(row=2, column=8, value="Stricter scorer (lower scores than the group).")
    ws.cell(row=4, column=7, value="Estimated Bias > 0:")
    ws.cell(row=4, column=8, value="More lenient scorer (higher scores than the group).")


def build_matrix_by_category(wb, d):
    """Build Matrix by Category & Applicant sheet. Skipped if no rubric labels exist."""
    ws = wb["Matrix by Category & Applicant"]

    label_list = list(d["labels"].keys())
    app_list = list(d["apps"].keys())

    if not label_list:
        ws.cell(row=1, column=1, value="No rubric categories found in this competition")
        return

    ws.cell(row=1, column=1, value="Sum of Rating")
    ws.cell(row=2, column=1, value="Reviewer/Applicant")

    for i, lbl in enumerate(label_list):
        ws.cell(row=2, column=2 + i, value=lbl)

    col_grand = 2 + len(label_list)
    ws.cell(row=2, column=col_grand, value="Grand Total")

    out_row = 3
    for idx, app in enumerate(app_list, 1):
        applicant = d["app_applicant"].get(app, "")
        ws.cell(row=out_row, column=1, value=make_app_key(idx, app, applicant))

        for i, lbl in enumerate(label_list):
            total = 0
            for rev in d["reviewers_by_app"].get(app, {}):
                total += d["cat_ratings"].get(app, {}).get(rev, {}).get(lbl, 0)
            ws.cell(row=out_row, column=2 + i, value=total)

        row_sum = sum(ws.cell(row=out_row, column=2 + i).value or 0 for i in range(len(label_list)))
        ws.cell(row=out_row, column=col_grand, value=row_sum)

        ws.cell(row=out_row, column=1).font = RED_FONT_BOLD

        for rev in d["reviewers_by_app"].get(app, {}):
            out_row += 1
            ws.cell(row=out_row, column=1, value=rev)

            for i, lbl in enumerate(label_list):
                val = d["cat_ratings"].get(app, {}).get(rev, {}).get(lbl, None)
                if val is not None:
                    ws.cell(row=out_row, column=2 + i, value=val)
                else:
                    ws.cell(row=out_row, column=2 + i, value="")

            rev_sum = sum(d["cat_ratings"].get(app, {}).get(rev, {}).get(lbl, 0) for lbl in label_list)
            ws.cell(row=out_row, column=col_grand, value=rev_sum)

        out_row += 1

    last_row = out_row - 1
    max_col = col_grand

    # Formatting
    ws.row_dimensions[1].hidden = True
    ws.row_dimensions[2].height = 87
    for r in range(3, last_row + 1):
        ws.row_dimensions[r].height = 87

    ws.column_dimensions["A"].width = 32
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    apply_header_row(ws, 2, max_col)
    apply_center_all(ws, 1, last_row, 1, max_col)


def build_reviewer_admin_comments(wb, d):
    """Build Reviewer Administrator Comments sheet."""
    ws = wb["Reviewer Administrator Comments"]

    if not d["has_admin"] or not d["admin_comments"]:
        ws.cell(row=1, column=1, value="No administrator comments found in this competition")
        ws.column_dimensions["A"].width = 60
        return

    app_list = list(d["apps"].keys())
    ws.cell(row=1, column=1, value="Row Labels")

    out_row = 2
    for reviewer, app_comments in d["admin_comments"].items():
        ws.cell(row=out_row, column=1, value=reviewer)
        ws.cell(row=out_row, column=1).fill = BLACK_FILL
        ws.cell(row=out_row, column=1).font = WHITE_FONT
        out_row += 1

        for app in app_list:
            idx = d["apps"][app]
            applicant = d["app_applicant"].get(app, "")

            ws.cell(row=out_row, column=1, value=idx)
            ws.cell(row=out_row, column=1).fill = GRAY_FILL
            out_row += 1

            ws.cell(row=out_row, column=1, value=" " + make_app_key(idx, app, applicant))
            ws.cell(row=out_row, column=1).fill = GRAY_FILL
            out_row += 1

            cmt = app_comments.get(app, "")
            if not cmt or not cmt.strip():
                cmt = "--"
            ws.cell(row=out_row, column=1, value=cmt)
            out_row += 1

    last_row = out_row - 1

    ws.column_dimensions["A"].width = 98
    for r in range(1, last_row + 1):
        cell = ws.cell(row=r, column=1)
        cell.alignment = LEFT_TOP_WRAP
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=11)


def build_reviewer_comments(wb, d):
    """Build Reviewer Comments sheet."""
    ws = wb["Reviewer Comments"]

    if not d["has_comments"] or not d["comments"]:
        ws.cell(row=1, column=1, value="No reviewer comments found in this competition")
        ws.column_dimensions["A"].width = 60
        return

    app_list = list(d["apps"].keys())
    ws.cell(row=1, column=1, value="Row Labels")

    out_row = 2
    for app in app_list:
        idx = d["apps"][app]
        applicant = d["app_applicant"].get(app, "")

        ws.cell(row=out_row, column=1, value=idx)
        ws.cell(row=out_row, column=1).fill = BLACK_FILL
        ws.cell(row=out_row, column=1).font = WHITE_FONT_BOLD
        out_row += 1

        ws.cell(row=out_row, column=1, value=applicant if applicant else "(no applicant name)")
        ws.cell(row=out_row, column=1).fill = BLACK_FILL
        ws.cell(row=out_row, column=1).font = WHITE_FONT_BOLD
        out_row += 1

        if app in d["comments"]:
            for reviewer in d["comments"][app]:
                ws.cell(row=out_row, column=1, value=reviewer)
                ws.cell(row=out_row, column=1).fill = LIGHT_GRAY_FILL
                ws.cell(row=out_row, column=1).font = BLACK_FONT
                out_row += 1

                for label, comment in d["comments"][app][reviewer]:
                    ws.cell(row=out_row, column=1, value=f"{label} - {comment}")
                    ws.cell(row=out_row, column=1).fill = WHITE_FILL
                    ws.cell(row=out_row, column=1).font = BLACK_FONT
                    out_row += 1

    last_row = out_row - 1

    ws.column_dimensions["A"].width = 126
    for r in range(1, last_row + 1):
        cell = ws.cell(row=r, column=1)
        cell.alignment = LEFT_TOP_WRAP
        cell.border = THIN_BORDER
        if not cell.font:
            cell.font = Font(name="Calibri", size=11)


# ============================================================
# MAIN
# ============================================================

def main():
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        candidates = []
        for f in os.listdir("."):
            if f.lower().endswith((".xlsx", ".xlsb", ".csv")) and "macroresult" not in f.lower():
                candidates.append(f)

        if len(candidates) == 1:
            input_file = candidates[0]
            print(f"Auto-detected input file: {input_file}")
        elif len(candidates) > 1:
            print("Multiple data files found. Please specify which one:")
            for i, f in enumerate(candidates, 1):
                print(f"  {i}. {f}")
            choice = input("Enter number: ").strip()
            try:
                input_file = candidates[int(choice) - 1]
            except (ValueError, IndexError):
                print("Invalid choice.")
                sys.exit(1)
        else:
            print("Usage: python3 rev_rating.py <input_file.xlsx>")
            print("\nAccepted formats: .xlsx, .xlsb, .csv")
            print("The file should be an InfoReady Application Reviews export")
            print("containing a 'Reviewer Details' sheet.")
            sys.exit(1)

    if not os.path.exists(input_file):
        print(f"ERROR: File not found: {input_file}")
        sys.exit(1)

    print(f"Reading: {input_file}")
    headers, data = read_reviewer_details(input_file)

    if not data:
        print("ERROR: No data found. Check that the file contains a Reviewer Details sheet with data rows.")
        sys.exit(1)

    print(f"  Found {len(data)} rows, {len(headers)} columns")

    # Process
    d = process_data(headers, data)
    print(f"\n  Summary:")
    print(f"    Applications: {len(d['apps'])}")
    print(f"    Reviewers: {len(d['reviewers'])}")
    print(f"    Rating categories: {len(d['labels'])}")
    if not d['has_labels']:
        print(f"    Note: No rubric categories found. Matrix by Category sheet will be empty.")
    if not d['has_comments']:
        print(f"    Note: No applicant comments column found. Reviewer Comments sheet will be empty.")
    if not d['has_admin']:
        print(f"    Note: No admin comments column found. Admin Comments sheet will be empty.")

    # Create output workbook
    wb = openpyxl.Workbook()
    wb.active.title = "Matrix Final"
    wb.create_sheet("Reviewer Summary")
    wb.create_sheet("Matrix by Category & Applicant")
    wb.create_sheet("Reviewer Administrator Comments")
    wb.create_sheet("Reviewer Comments")

    # Build all sheets
    print("\nBuilding sheets...")
    mf_info = build_matrix_final(wb, d)
    build_reviewer_summary(wb, mf_info)
    build_matrix_by_category(wb, d)
    build_reviewer_admin_comments(wb, d)
    build_reviewer_comments(wb, d)

    # Save alongside input file, or fallback to current directory
    input_dir = os.path.dirname(os.path.abspath(input_file))
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_name = f"MacroResult_{timestamp}.xlsx"
    output_path = os.path.join(input_dir, output_name)

    if not os.access(input_dir, os.W_OK):
        output_path = os.path.join(os.getcwd(), output_name)

    print(f"\nSaving: {output_path}")
    wb.save(output_path)
    print(f"Done! Output saved to:\n  {output_path}")


if __name__ == "__main__":
    main()
